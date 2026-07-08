"""
Amazon Product Upload Generator — single GROUPID, on-demand (server helper).

A faithful port of the standalone C:\\scripts\\amz-product\\amz_upload.py, adapted so the
bcweb Express server can invoke it for ONE product on demand (the "Amazon upload file"
button on the Add / Modify screen). The heavy lifting is identical: it injects rows into a
copy of the VBA-enabled SHOES.xlsm template — that embedded macro/settings payload is what
makes Amazon Seller Central accept the .xlsm, and no Node Excel library preserves it, which
is why this stays in Python (openpyxl keep_vba=True).

Differences from the standalone script (deliberate):
  - Reads ONE groupid + the output path from argv, instead of groupids.txt / a fixed name.
  - Loads DB creds from the SERVER's .env (bcweb-server/.env), not C:\\scripts\\.env.
  - Prints a machine-readable JSON summary to stdout for the Node route; on any failure it
    prints {"error": "<CODE>", "message": "..."} and exits non-zero. No Downloads copy.
  - DB write behaviour is IDENTICAL ("mirror the script exactly", owner-confirmed): it stamps
    skumap (sku, status='1', updated) for the variants written, and skips variants already in
    the amzfeed table so they are not re-listed.

Usage:
    python amz_upload_single.py <GROUPID> <OUTPUT_XLSM_PATH>

Exit codes: 0 = success (JSON summary on stdout); non-zero = failure (JSON error on stdout).
"""

import os
import sys
import json
import shutil
from datetime import datetime

import openpyxl
import psycopg2
from dotenv import load_dotenv

# Paths — the template lives next to this script; .env is the server's (two dirs up).
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "SHOES.xlsm")
SERVER_ENV = os.path.join(SCRIPT_DIR, "..", "..", ".env")

# Column positions in SHOES.xlsm Template sheet (openpyxl is 1-based) — copied verbatim from amz_upload.py.
COL_SKU = 1            # A
COL_PRODUCT_TYPE = 2   # B
COL_LISTING_ACTION = 3 # C
COL_BRAND = 8          # H
COL_PROD_ID_TYPE = 9   # I
COL_PROD_ID = 10       # J
COL_CONDITION = 168    # FL
COL_LIST_PRICE = 170   # FN
COL_FULFILLMENT = 193  # GK
COL_PRICE = 198        # GP
COL_BATTERIES_REQ = 267  # JG - "Are batteries required?"

# Data rows start at row 7 in the template (rows 1-6 are settings/headers/examples).
DATA_START_ROW = 7


def fail(code, message):
    """Emit a JSON error envelope on stdout and exit non-zero for the Node route to map to a return_code."""
    print(json.dumps({"error": code, "message": message}))
    sys.exit(1)


def get_db_connection():
    """Connect to PostgreSQL using the server's .env credentials."""
    load_dotenv(SERVER_ENV)
    try:
        return psycopg2.connect(
            host=os.environ["DB_HOST"],
            port=os.environ.get("DB_PORT", "5432"),
            dbname=os.environ["DB_NAME"],
            user=os.environ["DB_USER"],
            password=os.environ["DB_PASSWORD"],
        )
    except KeyError as e:
        fail("DB_CONFIG", f"Missing DB env var: {e}")
    except Exception as e:
        fail("DB_CONNECT", f"Could not connect to the database: {e}")


def fetch_existing_codes(conn, groupid):
    """Return set of codes already in amzfeed for this GROUPID (so we don't re-list them)."""
    with conn.cursor() as cur:
        cur.execute("SELECT code FROM amzfeed WHERE groupid = %s", (groupid,))
        return {row[0] for row in cur.fetchall()}


def fetch_product_data(conn, groupid):
    """Fetch brand/rrp from skusummary and per-size variant data from skumap for ONE groupid."""
    rows = []
    skipped = 0
    yymm = datetime.now().strftime("%y%m")
    existing_codes = fetch_existing_codes(conn, groupid)

    with conn.cursor() as cur:
        cur.execute("SELECT brand, rrp FROM skusummary WHERE groupid = %s", (groupid,))
        summary = cur.fetchone()
        if not summary:
            fail("NOT_FOUND", f"GROUPID '{groupid}' not found in skusummary")

        brand, rrp = summary
        if not brand:
            fail("NO_BRAND", f"No brand set for GROUPID '{groupid}'")

        try:
            rrp_val = float(rrp) if rrp else 0.0
        except ValueError:
            fail("INVALID_RRP", f"Invalid RRP '{rrp}' for GROUPID '{groupid}'")

        cur.execute(
            """SELECT sku, code, ean
               FROM skumap
               WHERE groupid = %s AND deleted = 0
               ORDER BY code""",
            (groupid,),
        )
        variants = cur.fetchall()

        if not variants:
            fail("NO_SIZES", f"No active variants for GROUPID '{groupid}'")

        for sku, code, ean in variants:
            # Skip variants already on Amazon (mirror the standalone script).
            if code in existing_codes:
                skipped += 1
                continue

            amz_sku = sku.strip() if (sku and sku.strip()) else f"{code}-{yymm}"

            clean_ean = ean.rstrip("B") if ean else ""
            if not clean_ean:
                skipped += 1
                continue

            rows.append({
                "sku": amz_sku,
                "code": code,
                "brand": brand,
                "ean": clean_ean,
                "rrp": rrp_val,
            })

    return rows, skipped


def update_skumap(conn, rows):
    """Stamp skumap (sku, status='1', updated) for the written variants — identical to the standalone script."""
    today = datetime.now().strftime("%Y-%m-%d")
    updated = 0
    with conn.cursor() as cur:
        for row in rows:
            cur.execute(
                """UPDATE skumap
                   SET sku = %s, status = '1', updated = %s
                   WHERE UPPER(code) = UPPER(%s)""",
                (row["sku"], today, row["code"]),
            )
            updated += cur.rowcount
    conn.commit()
    return updated


def generate_upload(rows, output_path):
    """Copy the SHOES.xlsm template and inject the product rows, saving to output_path."""
    shutil.copy2(TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path, keep_vba=True)
    ws = wb["Template"]

    # Clear any example/left-over data rows from the template.
    for r in range(DATA_START_ROW, DATA_START_ROW + 1000):
        if ws.cell(row=r, column=COL_SKU).value is None:
            break
        for col in [COL_SKU, COL_PRODUCT_TYPE, COL_LISTING_ACTION, COL_BRAND,
                    COL_PROD_ID_TYPE, COL_PROD_ID, COL_CONDITION,
                    COL_LIST_PRICE, COL_FULFILLMENT, COL_PRICE, COL_BATTERIES_REQ]:
            ws.cell(row=r, column=col, value=None)

    for i, row in enumerate(rows):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=COL_SKU, value=row["sku"])
        ws.cell(row=r, column=COL_PRODUCT_TYPE, value="SHOES")
        ws.cell(row=r, column=COL_LISTING_ACTION, value="partial_update")
        ws.cell(row=r, column=COL_BRAND, value=row["brand"])
        ws.cell(row=r, column=COL_PROD_ID_TYPE, value="EAN")
        ws.cell(row=r, column=COL_PROD_ID, value=row["ean"])
        ws.cell(row=r, column=COL_CONDITION, value="New")
        ws.cell(row=r, column=COL_LIST_PRICE, value=row["rrp"])
        ws.cell(row=r, column=COL_FULFILLMENT, value="Fulfilment by Merchant (Default)")
        ws.cell(row=r, column=COL_PRICE, value=row["rrp"])
        ws.cell(row=r, column=COL_BATTERIES_REQ, value="No")

    wb.save(output_path)


def main():
    if len(sys.argv) != 3:
        fail("BAD_ARGS", "Usage: amz_upload_single.py <GROUPID> <OUTPUT_XLSM_PATH>")

    groupid = sys.argv[1].strip().upper()
    output_path = sys.argv[2]
    if not groupid:
        fail("BAD_ARGS", "GROUPID is empty")
    if not os.path.exists(TEMPLATE_PATH):
        fail("NO_TEMPLATE", f"Template not found at {TEMPLATE_PATH}")

    conn = get_db_connection()
    try:
        rows, skipped = fetch_product_data(conn, groupid)
        if not rows:
            # Everything was skipped (already on Amazon, or no EANs) — nothing to upload.
            fail("NO_ROWS", "No variants to write (all already on Amazon, or missing EANs)")
        generate_upload(rows, output_path)
        updated = update_skumap(conn, rows)
    finally:
        conn.close()

    print(json.dumps({
        "groupid": groupid,
        "variants": len(rows),
        "skipped": skipped,
        "skumapUpdated": updated,
        "brand": rows[0]["brand"],
    }))


if __name__ == "__main__":
    main()
