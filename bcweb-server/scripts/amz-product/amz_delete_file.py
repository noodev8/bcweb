"""
Amazon Listing DELETE file generator (server helper).

Builds the Seller Central .xlsm that REMOVES listings — the mirror image of amz_upload_single.py.
It exists for one job: the "on Amazon, unknown to us" bucket on the Update Amazon screen. Those are
SKUs Amazon still lists but we hold no product record for; the zero-stock ones are dead listings and
the only way to clear them is a Listings feed with the Listing Action set to delete.

Why Python and not Node (same reason as amz_upload_single.py): the file is a macro-enabled .xlsm and
Amazon accepts it only because of the settings payload in row 1 of the Template sheet. No Node Excel
library round-trips that intact; openpyxl (keep_vba=True) does.

TEMPLATE REUSE — deliberate. This writes into the SAME SHOES.xlsm that amz_upload_single.py uses,
which lives in this repo (bcweb-server/scripts/amz-product/SHOES.xlsm) and is git-tracked. A delete
row needs only three of its 348 columns, and they sit in identical positions in every Seller Central
Listings template for SHOES/UK, so a second template would be 1.1 MB of duplicate that ages on its
own. Refreshing SHOES.xlsm now fixes both flows at once.

NO DATABASE. Unlike amz_upload_single.py this touches nothing — it does not connect to Postgres and
writes no rows. It is a pure formatter: SKUs in, file out. That is the point. The SKUs it is given
are by definition unknown to our database, so there is nothing here to stamp; and a delete is applied
by Amazon when the operator uploads the file, not by us. Keeping it read-nothing/write-nothing means
clicking the button can never leave our data in a state that assumes an upload that never happened.

Input is a JSON array of SKU strings on STDIN, not argv: the list can run to hundreds of entries and
Windows caps a command line at ~32k characters.

Usage:
    echo '["SKU-1","SKU-2"]' | python amz_delete_file.py <OUTPUT_XLSM_PATH>

Exit codes: 0 = success (JSON summary on stdout); non-zero = failure (JSON error on stdout).
"""

import os
import sys
import json
import shutil

import openpyxl

# The template lives next to this script — same copy amz_upload_single.py uses.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "SHOES.xlsm")

# Column positions in the SHOES.xlsm Template sheet (openpyxl is 1-based). A delete needs only these
# three; every other column must stay EMPTY or Amazon treats the row as a listing update instead.
#   A = contribution_sku#1.value   B = product_type#1.value   C = ::record_action
COL_SKU = 1
COL_PRODUCT_TYPE = 2
COL_LISTING_ACTION = 3

# Data rows start at row 7 (1 = settings blob, 3-4 = headings/labels, 5 = attribute names, 6 = examples).
DATA_START_ROW = 7

# The product type the template was generated for. Not a free choice — it must match the template's
# own settings payload, and SHOES.xlsm is the SHOES/UK template.
PRODUCT_TYPE = "SHOES"

# ::record_action value for a removal. The raw enum, not the "Delete" dropdown label, to match
# amz_upload_single.py which writes "partial_update" the same way and is accepted in production.
RECORD_ACTION_DELETE = "delete"

# Safety rail. The unknown-SKU bucket is normally a handful to a few hundred; a four-figure delete file
# means something upstream is wrong (a truncated or mis-parsed inventory report), and refusing is a lot
# cheaper than an operator uploading a file that de-lists the catalogue.
MAX_SKUS = 1000


def fail(code, message):
    """Emit a JSON error envelope on stdout and exit non-zero for the Node route to map to a return_code."""
    print(json.dumps({"error": code, "message": message}))
    sys.exit(1)


def generate_delete_file(skus, output_path):
    """Copy the template and write one delete row per SKU, saving to output_path."""
    shutil.copy2(TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path, keep_vba=True)
    ws = wb["Template"]

    # Clear any example/left-over data rows the template shipped with, so the file contains our rows
    # and nothing else. Mirrors amz_upload_single.py; only the three columns we write can be dirty.
    for r in range(DATA_START_ROW, DATA_START_ROW + 1000):
        if ws.cell(row=r, column=COL_SKU).value is None:
            break
        for col in (COL_SKU, COL_PRODUCT_TYPE, COL_LISTING_ACTION):
            ws.cell(row=r, column=col, value=None)

    for i, sku in enumerate(skus):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=COL_SKU, value=sku)
        ws.cell(row=r, column=COL_PRODUCT_TYPE, value=PRODUCT_TYPE)
        ws.cell(row=r, column=COL_LISTING_ACTION, value=RECORD_ACTION_DELETE)

    wb.save(output_path)


def main():
    if len(sys.argv) != 2:
        fail("BAD_ARGS", "Usage: amz_delete_file.py <OUTPUT_XLSM_PATH>  (SKU JSON array on stdin)")

    output_path = sys.argv[1]
    if not os.path.exists(TEMPLATE_PATH):
        fail("NO_TEMPLATE", f"Template not found at {TEMPLATE_PATH}")

    try:
        payload = json.load(sys.stdin)
    except Exception as e:
        fail("BAD_ARGS", f"Could not read the SKU list from stdin: {e}")

    if not isinstance(payload, list):
        fail("BAD_ARGS", "Expected a JSON array of SKU strings on stdin")

    # De-duplicate while preserving order: two rows for one SKU is a rejected feed, and the caller
    # builds this list from a report we do not control.
    seen = set()
    skus = []
    for raw in payload:
        sku = str(raw).strip()
        if not sku or sku in seen:
            continue
        seen.add(sku)
        skus.append(sku)

    if not skus:
        fail("NO_ROWS", "No SKUs to delete")
    if len(skus) > MAX_SKUS:
        fail("TOO_MANY", f"{len(skus)} SKUs exceeds the {MAX_SKUS} safety limit for one delete file")

    generate_delete_file(skus, output_path)
    print(json.dumps({"skus": len(skus)}))


if __name__ == "__main__":
    main()
